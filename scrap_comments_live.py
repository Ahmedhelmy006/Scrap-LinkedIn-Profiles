from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import json
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

class LinkedInLiveCommentScraper:
    def __init__(self, cookies_file=None):
        self.cookies_file = cookies_file
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            executable_path=r"C:\Program Files\Google\Chrome\Application\chrome.exe",  # Provide the path to the Chrome executable
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-infobars',
                '--disable-dev-shm-usage',
                '--disable-extensions',
                '--disable-gpu',
            ]
)
        self.context = self.browser.new_context()
        if cookies_file:
            with open(cookies_file, 'r') as file:
                cookies = json.load(file)
                self.context.add_cookies(cookies)
        self.page = self.context.new_page()

    def load_event_page(self, url, sleep_time):
        self.page.goto(url)
        print(f"Sleeping for {sleep_time} seconds to allow comments to load...")
        time.sleep(sleep_time)


    def parse_comments(self):
        html = self.page.content()
        soup = BeautifulSoup(html, 'html.parser')

        comment_articles = soup.find_all('article', class_='comments-comment-item')
        commentators = {}

        for article in comment_articles:
            try:
                # Extract Profile URL
                profile_link_tag = article.find('a', class_='comments-post-meta__actor-link')
                profile_url = profile_link_tag['href'] if profile_link_tag else "N/A"

                # Extract Commentator Name
                name_tag = article.find('span', class_='comments-post-meta__name-text')
                name = name_tag.get_text(strip=True) if name_tag else "N/A"

                # Extract Comment Text
                comment_tag = article.find('span', class_='comments-comment-item__main-content')
                comment = comment_tag.get_text(strip=True) if comment_tag else "N/A"

                if name not in commentators:
                    commentators[name] = {
                        "Commentator Name": name,
                        "Commentator Profile URL": profile_url,
                        "Number of comments": 0,
                        "Comments": []
                    }
                commentators[name]["Number of comments"] += 1
                commentators[name]["Comments"].append(comment)
            except Exception as e:
                print(f"Error while parsing a comment: {e}")

        return commentators



    def close(self):
        self.context.close()
        self.browser.close()
        self.playwright.stop()


def save_to_xlsx(data, file_name):
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comments"

        # Headers
        sheet.append(["Commentator Name", "Commentator Profile URL", "Number of Comments", "Comments"])

        for item in data:
            comments_bulleted = '\n'.join([f"\u2022 {comment}" for comment in item["comments"]])
            sheet.append([
                item["Commentator Name"],
                item["Commentator Profile URL"],
                item["Number of comments"],
                comments_bulleted
            ])

        # Align the comments column to the top-left
        for row in sheet.iter_rows(min_row=2, max_col=4):
            for cell in row:
                if cell.column == 4:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        workbook.save(file_name)
        print(f"Data saved to {file_name}")
    except Exception as e:
        print(f"Error saving data to Excel: {e}")


def main():
    cookies_file = r"Input Files/cookies.json"
    linkedin_live_event_url = "https://www.linkedin.com/video/live/urn:li:ugcPost:7270816698314768384/"
    sleep_time = 720  # Customize sleep time for real-time comment collection

    scraper = LinkedInLiveCommentScraper(cookies_file)
    scraper.load_event_page(linkedin_live_event_url, sleep_time)

    # Parse comments
    commentators = scraper.parse_comments()
    scraper.close()

    # Convert data to list format
    commentators_list = list(commentators.values())

    # Save to JSON
    with open("live_comments.json", "w", encoding="utf-8") as file:
        json.dump(commentators_list, file, indent=4)

    # Save to XLSX
    save_to_xlsx(commentators_list, "live_comments.xlsx")

    # Print statistics
    unique_commentators = len(commentators_list)
    total_comments = sum(c["Number of comments"] for c in commentators_list)

    print(f"Total unique commentators: {unique_commentators}")
    print(f"Total comments: {total_comments}")


if __name__ == "__main__":
    main()

