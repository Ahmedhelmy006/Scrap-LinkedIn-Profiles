from playwright.sync_api import sync_playwright
import pandas as pd
from openpyxl import Workbook, load_workbook
import time
import random
import json
from bs4 import BeautifulSoup

class LinkedInScraper:
    def __init__(self, cookies_file=None):
        self.cookies_file = cookies_file
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-infobars',
                '--disable-dev-shm-usage',
                '--disable-extensions',
                '--disable-gpu',
            ]
        )
        self.context = self.browser.new_context()
        if cookies_file:
            with open(cookies_file, 'r') as file:
                cookies = json.load(file)
                self.context.add_cookies(cookies)
        self.page = self.context.new_page()

    def load_profile_page(self, profile_url):
        self.page.goto(profile_url)
        time.sleep(5)
        self.scroll_to_bottom()

    def scroll_to_bottom(self):
        while True:
            previous_height = self.page.evaluate("document.body.scrollHeight")
            self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(2)
            current_height = self.page.evaluate("document.body.scrollHeight")
            if current_height == previous_height:
                break

    def parse_experience_block(self, html_snippet):
        soup = BeautifulSoup(html_snippet, "html.parser")
        all_experiences = []

        experience_items = soup.select("li.artdeco-list__item")
        for item in experience_items:
            job_title_elem = item.select_one("div.display-flex.align-items-center.mr1.t-bold span[aria-hidden='true']")
            job_title = job_title_elem.get_text(strip=True) if job_title_elem else "N/A"

            company_elem = item.select_one("span.t-14.t-normal span[aria-hidden='true']")
            company_name = company_elem.get_text(strip=True) if company_elem else "N/A"

            dl_elems = item.select("span.t-14.t-normal.t-black--light span[aria-hidden='true']")
            date_range = dl_elems[0].get_text(strip=True) if len(dl_elems) > 0 else "N/A"
            location = dl_elems[1].get_text(strip=True) if len(dl_elems) > 1 else "N/A"

            if job_title == "N/A" and company_name == "N/A" and date_range == "N/A" and location == "N/A":
                continue

            all_experiences.append(f"{job_title}\n{company_name}\n{date_range}\n{location}")

        return "\n\n".join(all_experiences) if all_experiences else "No experience information found."

    def parse_education_block(self, html_snippet):
        soup = BeautifulSoup(html_snippet, "html.parser")
        all_education = []
        education_items = soup.select("li.artdeco-list__item, li.pvs-list__item")
        for item in education_items:
            school_elem = item.select_one("span.mr1.hoverable-link-text")
            school_name = school_elem.get_text(strip=True) if school_elem else "N/A"

            degree_elem = item.select_one("span.t-14.t-normal")
            degree_name = degree_elem.get_text(strip=True) if degree_elem else "N/A"

            date_elem = item.select_one("span.t-14.t-normal.t-black--light")
            date_range = date_elem.get_text(strip=True) if date_elem else "N/A"

            if school_name == "N/A" and degree_name == "N/A" and date_range == "N/A":
                continue
            all_education.append(f"{school_name}\n{degree_name}\n{date_range}")

        return "\n\n".join(all_education) if all_education else "No education information found."

    def parse_about_section(self, soup):
        about_heading = soup.find(lambda tag: tag.name == "h2" and "about" in tag.get_text(strip=True).lower())
        if not about_heading:
            return "No about information found."

        heading_parent = about_heading.find_parent("div", class_="pvs-header__top-container--no-stack")
        if not heading_parent:
            return "No about information found."

        about_container = heading_parent.parent.find_next_sibling("div")
        if not about_container:
            return "No about information found."

        about_text = about_container.get_text(strip=True)
        if not about_text:
            return "No about information found."
        return about_text

    def parse_profile(self):
        html = self.page.content()
        soup = BeautifulSoup(html, 'html.parser')

        # Name
        name_elem = soup.select_one("h1")
        profile_name = self.clean_text(name_elem)

        # Job
        job_elem = soup.select_one("div.text-body-medium")
        job_title = self.clean_text(job_elem)

        # Location
        location = self.get_location(soup)

        # Followers
        followers_count = self.get_followers(soup)

        # Experience
        experience_heading = soup.find(lambda tag: tag.name == "h2" and "experience" in tag.get_text(strip=True).lower())
        if experience_heading:
            heading_parent = experience_heading.find_parent("div", class_="pvs-header__top-container--no-stack")
            if heading_parent:
                experience_container = heading_parent.parent.find_next_sibling("div")
                if experience_container:
                    experience_html = str(experience_container)
                    experience = self.parse_experience_block(experience_html)
                else:
                    experience = "No experience information found."
            else:
                experience = "No experience information found."
        else:
            experience = "No experience information found."

        # Education
        education_heading = soup.find(lambda tag: tag.name == "h2" and "education" in tag.get_text(strip=True).lower())
        if education_heading:
            edu_heading_parent = education_heading.find_parent("div", class_="pvs-header__top-container--no-stack")
            if edu_heading_parent:
                education_container = edu_heading_parent.parent.find_next_sibling("div")
                if education_container:
                    education_html = str(education_container)
                    education = self.parse_education_block(education_html)
                else:
                    education = "No education information found."
            else:
                education = "No education information found."
        else:
            education = "No education information found."

        # About
        about = self.parse_about_section(soup)

        return {
            "Profile Link": self.page.url,
            "Name": profile_name,
            "Job": job_title,
            "Location": location,
            "Followers": followers_count,
            "Education": education,
            "Experience": experience,
            "About": about
        }

    def get_location(self, soup):
        element = soup.select_one("span.text-body-small.inline.t-black--light.break-words")
        return self.clean_text(element)

    def get_followers(self, soup):
        element = soup.select_one("ul span.t-bold")
        return self.clean_text(element)

    def clean_text(self, element):
        return element.get_text(strip=True) if element else "N/A"

    def close(self):
        self.context.close()
        self.browser.close()
        self.playwright.stop()

def scrape_profiles_from_excel(file_path, output_file, cookies_file=None, start_row=1, end_row=None):
    profile_urls_df = pd.read_excel(file_path)
    profile_urls = profile_urls_df['Profile Link'].tolist()

    if end_row is None:
        end_row = len(profile_urls)

    try:
        workbook = load_workbook(output_file)
    except FileNotFoundError:
        workbook = Workbook()

    if "Likes" not in workbook.sheetnames:
        worksheet = workbook.create_sheet("Comments")
        worksheet.append(["Profile Link", "Name", "Job", "Location", "Followers", "Education", "Experience", "About"])
    else:
        worksheet = workbook["Comments"]

    scraper = LinkedInScraper(cookies_file)

    for index, profile_url in enumerate(profile_urls[start_row - 1:end_row], start=start_row):
        print(f"Scraping profile {index}/{end_row}: {profile_url}")
        try:
            scraper.load_profile_page(profile_url)
            profile_info = scraper.parse_profile()
            worksheet.append([
                profile_info["Profile Link"],
                profile_info["Name"],
                profile_info["Job"],
                profile_info["Location"],
                profile_info["Followers"],
                profile_info["Education"],
                profile_info["Experience"],
                profile_info["About"]
            ])
            workbook.save(output_file)
            time.sleep(random.randint(2, 5))
        except Exception as e:
            print(f"Error scraping profile {profile_url}: {e}")

    scraper.close()
    print("All profiles scraped successfully.")

if __name__ == "__main__":
    input_file = 'comments_pt2_url.xlsx'
    output_file = 'comments_combined_final.xlsx'
    cookies_file = r'Input Files\cookies.json'
    start_row = 1
    end_row = 85
    scrape_profiles_from_excel(input_file, output_file, cookies_file, start_row=start_row, end_row=end_row)
