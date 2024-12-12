from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import json
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

class LinkedInCommentScraper:
    def __init__(self, cookies_file=None):
        self.cookies_file = cookies_file
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-infobars',
                '--disable-dev-shm-usage',
                '--disable-extensions',
                '--disable-gpu',
            ]
        )
        self.context = self.browser.new_context()
        if cookies_file:
            with open(cookies_file, 'r') as file:
                cookies = json.load(file)
                self.context.add_cookies(cookies)
        self.page = self.context.new_page()

    def load_comments_page(self, url):
        self.page.goto(url)
        time.sleep(5)

        while True:
            try:
                load_more_button = self.page.locator('button.comments-comments-list__load-more-comments-button--cr')
                if load_more_button.is_visible():
                    load_more_button.click()
                    time.sleep(2)
                else:
                    break
            except Exception as e:
                print(f"Error while clicking 'Load more comments': {e}")
                break

    def parse_comments(self):
        html = self.page.content()
        soup = BeautifulSoup(html, 'html.parser')

        comment_articles = soup.find_all('article', class_='comments-comment-entity')
        commentators = {}

        for article in comment_articles:
            try:
                profile_link_tag = article.find('a', class_='comments-comment-meta__image-link')
                profile_url = profile_link_tag['href'] if profile_link_tag else "N/A"

                name_tag = article.find('span', class_='comments-comment-meta__description-title')
                name = name_tag.get_text(strip=True) if name_tag else "N/A"

                comment_tag = article.find('span', class_='comments-comment-item__main-content')
                comment = comment_tag.get_text(strip=True) if comment_tag else "N/A"

                if name not in commentators:
                    commentators[name] = {
                        "Commentator Name": name,
                        "Commentator Profile URL": profile_url,
                        "Number of comments": 0,
                        "comments": []
                    }
                commentators[name]["Number of comments"] += 1
                commentators[name]["comments"].append(comment)
            except Exception as e:
                print(f"Error while parsing a comment: {e}")

        return commentators

    def close(self):
        self.context.close()
        self.browser.close()
        self.playwright.stop()


def save_to_xlsx(data, file_name):
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comments"

        # Headers
        sheet.append(["Commentator Name", "Commentator Profile URL", "Number of Comments", "Comments"])

        for item in data:
            comments_bulleted = '\n'.join([f"\u2022 {comment}" for comment in item["comments"]])
            sheet.append([
                item["Commentator Name"],
                item["Commentator Profile URL"],
                item["Number of comments"],
                comments_bulleted
            ])

        # Align the comments column to the top-left
        for row in sheet.iter_rows(min_row=2, max_col=4):
            for cell in row:
                if cell.column == 4:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        workbook.save(file_name)
        print(f"Data saved to {file_name}")
    except Exception as e:
        print(f"Error saving data to Excel: {e}")


def main():
    cookies_file = r"Input Files/cookies.json"
    linkedin_post_url = "https://www.linkedin.com/video/live/urn:li:ugcPost:7270816698314768384/"

    scraper = LinkedInCommentScraper(cookies_file)
    scraper.load_comments_page(linkedin_post_url)

    commentators = scraper.parse_comments()
    scraper.close()

    commentators_list = list(commentators.values())

    # Save to JSON
    with open("comments.json", "w", encoding="utf-8") as file:
        json.dump(commentators_list, file, indent=4)

    # Save to XLSX
    save_to_xlsx(commentators_list, "comments.xlsx")

    # Statistics
    unique_commentators = len(commentators_list)
    total_comments = sum(c["Number of comments"] for c in commentators_list)

    print(f"Total unique commentators: {unique_commentators}")
    print(f"Total comments: {total_comments}")


if __name__ == "__main__":
    main()